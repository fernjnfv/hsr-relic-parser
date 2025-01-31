import re
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import *
import os

def universal_parser(fact):
    superhero = re.findall(r"({\(*slug(.|\n)*?\)})", fact)
    all_heroes = dict()
    for i in range(len(superhero)):
        tmp = superhero[i][0]
        print()
        part = re.search(r"&& \(((.|\n)*?)\)}", tmp)
        hero_info = part.group(1)
        mrm = BeautifulSoup(hero_info, 'html.parser')
        #mrm = soup.find('div', attrs={'classname': 'detailed-cones moc extra planar'})
        mrm2 = mrm.find('div', attrs={'classname': '{`single-cone'})
        if mrm2 is not None:
            if mrm2.find('p').text != '1':
                print("error")
                break
            mytype = mrm2.find('div', attrs={'classname': 'double-set'})
        else:
            mytype = mrm.find('ul', attrs={'classname': 'with-sets'})
        relic_arr = []
        if mytype is not None:
            relic = mytype.findAll('hsrrelicset')
            for j in relic:
                relic_arr += [j.get("name")]
                print(j.get("name"))
            if len(relic) == 1:
                another_relic = mrm.find('ul', attrs={'classname': 'with-sets'})
                for j in another_relic.findAll('hsrrelicset'):
                    relic_arr += [j.get("name")+'*']
                    print(j.get('name'))
        else:
            relic = mrm2.find('hsrrelicset')
            relic_arr = [relic.get("name")]
            print(relic.get('name'))

        trer = re.findall(r"slug === '(.+?)'", tmp)
        for j in trer:
            all_heroes[j] = relic_arr
            print(j)
    return all_heroes

def planetary_parser(html_for_parse):
    relic = re.search(r"<h6>Best Planetary Sets</h6>(.|\n)*{/\* Sustain \*/}((.|\n)*)\r\n( )+<h6>Special Planetary Sets</h6>", html_for_parse)
    fact = relic.group(2)
    all_heroes = universal_parser(fact)
    return all_heroes

def relic_parse(html_for_parse):
    relic = re.search(r"{/\* Sustain \*/}((.|\n)*)\r\n( )+<h6>Best Planetary Sets</h6>", html_for_parse)
    fact = relic.group(1)
    all_heroes = universal_parser(fact)
    return all_heroes


def main_info():
    req = requests.get("https://www.prydwen.gg/component---src-dynamic-pages-hsr-character-dynamic-tsx-b94b7c55df6c3887f909.js.map")
    data = req.json()
    js_data = data["sourcesContent"][19]
    founded_html = re.search(r"return \(\r\n( )*<>\r\n((.|\r|\n)*)\r\n( )+</>\r\n( )+\);\r\n};", js_data)
    html_for_parse = founded_html.group(2)
    relic_heroes = relic_parse(html_for_parse)
    planetary_heroes = planetary_parser(html_for_parse)
    return {"relic": relic_heroes, "planetary": planetary_heroes}

def create_excel_from_dict_list(dict_list: list, output_filename: str, sheet_name='Sheet1'):
    # Создаем директорию, если она не существует
    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', output_filename)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Записываем данные из списка словарей в Excel
    if dict_list:
        header = list(dict_list[0].keys())
        ws.append(header)  # Записываем заголовки

        for row in dict_list:
            ws.append([row[col] for col in header])

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_style.border = border_style

    cell_style = NamedStyle(name='cell')
    cell_style.alignment = Alignment(horizontal='left', vertical='center')
    cell_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["E"].width = ws.column_dimensions["E"].width-17
    for i in range (2,100):
        ws.row_dimensions[i].height = 40

    # Сохраняем файл
    wb.save(filepath)
    return filepath

def str_stat_creation(relic):
    mystr = ""
    for i in relic:
        mystr += i["stat"]
        if i["sign"] is not None:
            mystr += i["sign"]
    return mystr

def parce_characters():
    heroes_info = main_info()
    relic_info = heroes_info["relic"]
    planar_info = heroes_info["planetary"]
    req = requests.get("https://www.prydwen.gg/page-data/sq/d/2408139295.json")
    req2 = requests.get("https://www.prydwen.gg/page-data/sq/d/2607493255.json")
    all_relic = req2.json()["data"]["allContentfulHsrRelics"]["nodes"]
    all_relic_img = dict()
    for relic in all_relic:
        tmp = relic["image"]["localFile"]["childImageSharp"]["gatsbyImageData"]["images"]["fallback"]["src"]
        all_relic_img[relic["name"]] = f"=_xlfn.IMAGE(\"https://www.prydwen.gg{tmp}\")"
    data = req.json()
    character_dict = dict()
    all_charecters_data1 = []
    all_charecters_data2 = []
    for character in data["data"]["allContentfulHsrCharacter"]["nodes"]:
        #print(character["slug"], character["name"], character["smallImage"]["localFile"]["childImageSharp"]["gatsbyImageData"]["images"]["fallback"]["src"])
        subreq = requests.get(f"https://www.prydwen.gg/page-data/star-rail/characters/{character["slug"]}/page-data.json")
        relic_character_info = relic_info.get(character["slug"], None)
        planar_character_info = planar_info.get(character["slug"], None)
        subdata = subreq.json()
        couter_cicle = 0
        for character_info in subdata["result"]["data"]["currentUnit"]["nodes"]:
            couter_cicle += 1
            temp_data1 = dict()
            temp_data2 = dict()
            temp_data1["name"] = character_info["name"]
            temp_data1["img"] = character_info["smallImage"]["localFile"]["childImageSharp"]["gatsbyImageData"]["images"]["fallback"]["src"]
            temp_data1["img"] = f"=_xlfn.IMAGE(\"https://www.prydwen.gg{temp_data1["img"]}\")"
            temp_data2["name"] = character_info["name"]
            temp_data2["img"] = temp_data1["img"]
            #print(temp_data["img"])
            ratings = character_info["ratings"]
            moc = ratings["moc"]
            pure = ratings["pure"]
            apo = ratings["apo"]
            if int(moc) <8 and int(pure) <8 and int(apo) <8:
                continue
            #print(character["slug"], relic_character_info)
            if character_info["buildData"] is None:
                #print(character_info["slug"], "no build data")
                temp_data1["body"] = "None"
                temp_data1["feet"] = "None"
                temp_data2["rope"] = "None"
                temp_data2["sphere"] = "None"
                #temp_data1["role"] = "None"
                #temp_data2["role"] = "None"
                temp_data1["substats"] = "None"
                temp_data2["substats"] = "None"
                temp_data1["relic_img"] = "None"
                temp_data2["relic_img"] = "None"
                if relic_character_info is None:
                    temp_data1["relic1"] = "None"
                if planar_character_info is None:
                    temp_data2["planars1"] = "None"

                #temp_data["relic2"] = "None"
                #temp_data["planars2"] = "None"
            else:
                for i in character_info["buildData"]:
                    #(character_info["slug"], i["name"])
                    #print(i)
                    temp_data1["body"] = str_stat_creation(i["body"])
                    temp_data1["feet"] = str_stat_creation(i["feet"])
                    temp_data2["rope"] = str_stat_creation(i["rope"])
                    temp_data2["sphere"] = str_stat_creation(i["sphere"])
                    #temp_data1["role"] = i["name"]
                    #temp_data2["role"] = i["name"]
                    temp_data1["substats"] = i["substats"]
                    temp_data2["substats"] = i["substats"]
                    temp_data1["relic_img"] = "None"
                    temp_data2["relic_img"] = "None"
                    if i["relics"] is not None and len(i["relics"]) >= 1:
                        if relic_character_info is None:
                            temp_data1["relic1"] = i["relics"][0]["relic"]
                            temp_data1["relic_img"] = all_relic_img[temp_data1["relic1"]]
                        if planar_character_info is None:
                            temp_data2["planars1"] = i["planars"][0]["planar"]
                            temp_data2["relic_img"] = all_relic_img[temp_data2["planars1"]]
                        # if len(i["relics"]) >= 2:
                        #     print(character_info["slug"])
                        #     temp_data["relic2"] = i["relics"][1]["relic"]
                        #     temp_data["planars2"] = i["planars"][1]["planar"]
                        # else:
                        #     temp_data["relic2"] = "None"
                        #     temp_data["planars2"] = "None"
                    else:
                        #print(character_info["slug"], i["name"], "no relic data")
                        if relic_character_info is None:
                            temp_data1["relic1"] = "None"
                        if planar_character_info is None:
                            temp_data2["planars1"] = "None"
                        # temp_data["relic2"] = "None"
                        # temp_data["planars2"] = "None"
                    if relic_character_info is None:
                        all_charecters_data1 += [temp_data1.copy()]
                    else:
                        for j in relic_character_info:
                            temp_data1["relic1"] = j
                            temp_data1["relic_img"] = all_relic_img[j[:-1] if j[-1] == "*" else j]
                            all_charecters_data1 += [temp_data1.copy()]
                    if planar_character_info is None:
                        all_charecters_data2 += [temp_data2.copy()]
                    else:
                        for j in planar_character_info:
                            temp_data2["planars1"] = j
                            temp_data2["relic_img"] = all_relic_img[j[:-1] if j[-1] == "*" else j]
                            print(temp_data2)
                            all_charecters_data2 += [temp_data2.copy()]
        if couter_cicle > 1:
            print("error")
            break
    print(1)
    print(all_charecters_data1)
    print(all_charecters_data2)
    #create_excel_from_dict_list(all_charecters_data1,"fake_users1.xlsx", 'Sheet1')
    #create_excel_from_dict_list(all_charecters_data2, "fake_users2.xlsx", 'Sheet1')

parce_characters()
